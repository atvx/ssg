import os
import platform
import shutil
import subprocess
from pathlib import Path
from pdf2image import convert_from_path
from openpyxl import load_workbook
from openpyxl.styles import Side, Border
import logging
import urllib.parse

logger = logging.getLogger(__name__)


def ensure_utf8_encoding():
    """确保系统环境使用UTF-8编码"""
    import locale
    try:
        # 设置环境变量强制使用UTF-8
        os.environ['LC_ALL'] = 'C.UTF-8'
        os.environ['LANG'] = 'C.UTF-8'
        os.environ['LC_CTYPE'] = 'C.UTF-8'
        
        # 设置Python的默认编码
        locale.setlocale(locale.LC_ALL, 'C.UTF-8')
        logger.info("已设置UTF-8编码环境")
    except Exception as e:
        logger.warning(f"设置UTF-8编码环境失败: {e}")


def normalize_filename(filename: str) -> str:
    """
    标准化文件名，确保中文字符正确编码
    
    Args:
        filename: 原始文件名
        
    Returns:
        str: 标准化后的文件名
    """
    try:
        # 确保文件名是正确的UTF-8编码
        if isinstance(filename, bytes):
            filename = filename.decode('utf-8')
        
        # 处理URL编码的中文字符
        if '%' in filename or '<' in filename:
            try:
                # 尝试URL解码
                filename = urllib.parse.unquote(filename, encoding='utf-8')
            except:
                pass
        
        # 处理十六进制编码的中文字符（如<e5><b8><82>）
        if '<' in filename and '>' in filename:
            import re
            # 查找所有<xx>格式的编码
            hex_pattern = r'<([a-fA-F0-9]{2})>'
            matches = re.findall(hex_pattern, filename)
            if matches:
                try:
                    # 将十六进制转换为字节，然后解码为UTF-8
                    byte_data = bytes([int(hex_code, 16) for hex_code in matches])
                    decoded_text = byte_data.decode('utf-8')
                    # 替换编码部分
                    filename = re.sub(r'<[a-fA-F0-9]{2}>', '', filename)
                    # 找到第一个编码的位置，插入解码后的文本
                    if '_' in filename:
                        parts = filename.split('_')
                        filename = f"{decoded_text}_{parts[-1]}"
                    else:
                        filename = decoded_text
                except Exception as e:
                    logger.warning(f"解码十六进制文件名失败: {e}")
        
        return filename.strip()
    except Exception as e:
        logger.error(f"标准化文件名失败: {e}")
        return filename


def set_excel_landscape_format(xlsx_path: str, sheetname: str = None):
    """
    设置Excel文件为横向打印格式
    
    Args:
        xlsx_path: Excel文件路径
        sheetname: 工作表名称，可选
    """
    try:
        wb = load_workbook(xlsx_path)
        ws = wb[sheetname] if sheetname else wb.active
        
        # 横向 & A4
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        
        # 一页宽（不限行数）- 确保表格在横向上适应一页
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        
        # 确保启用适应页面模式而不是百分比缩放
        ws.page_setup.fitToPage = True
        
        # 内容水平居中
        ws.print_options.horizontalCentered = True
        
        # 设置适当的页边距（单位：英寸）
        ws.page_margins.left = 0.25
        ws.page_margins.right = 0.25
        ws.page_margins.top = 0.5
        ws.page_margins.bottom = 0.5
        ws.page_margins.header = 0.3
        ws.page_margins.footer = 0.3
        
        # 自动设置打印区域（所有已用单元格）
        min_row = ws.min_row
        max_row = ws.max_row
        min_col = ws.min_column
        max_col = ws.max_column
        start_cell = ws.cell(row=min_row, column=min_col).coordinate
        end_cell = ws.cell(row=max_row, column=max_col).coordinate
        ws.print_area = f"{start_cell}:{end_cell}"
        
        # 修复标题区右侧边框加粗问题
        # 确保标题行使用整行宽度
        for col in range(1, max_col + 1):
            cell = ws.cell(row=1, column=col)
            if col == max_col:  # 最后一列
                cell.border = Border(
                    left=cell.border.left if cell.border else Side(style='thin'),
                    right=Side(style='medium'),
                    top=cell.border.top if cell.border else Side(style='thin'),
                    bottom=cell.border.bottom if cell.border else Side(style='thin')
                )
        
        wb.save(xlsx_path)
        logger.info(f"Excel文件格式设置完成: {xlsx_path}")
        
    except Exception as e:
        logger.error(f"设置Excel格式时出错: {str(e)}")
        raise


def find_soffice_executable() -> str:
    """
    动态寻找LibreOffice的soffice可执行文件
    
    Returns:
        str: soffice可执行文件路径
        
    Raises:
        FileNotFoundError: 未找到soffice可执行文件
    """
    sys_name = platform.system()
    candidates = []
    
    if sys_name == "Windows":
        candidates += [
            r"C:\Program Files\LibreOffice\program\soffice.exe",
            r"C:\Program Files (x86)\LibreOffice\program\soffice.exe",
        ]
    elif sys_name == "Darwin":  # macOS
        candidates += [
            "/Applications/LibreOffice.app/Contents/MacOS/soffice",
            shutil.which("soffice"),
        ]
    else:  # Linux
        candidates += [shutil.which("soffice"), shutil.which("libreoffice")]
    
    for candidate in candidates:
        if candidate and Path(candidate).exists():
            logger.info(f"找到LibreOffice可执行文件: {candidate}")
            return candidate
    
    raise FileNotFoundError("无法找到 LibreOffice 的 soffice，可执行文件未安装或不在 PATH")


def convert_xlsx_to_pdf(xlsx_path: Path, output_dir: Path = None) -> Path:
    """
    将Excel文件转换为PDF
    
    Args:
        xlsx_path: Excel文件路径
        output_dir: 输出目录，如果不指定则使用Excel文件所在目录
        
    Returns:
        Path: 生成的PDF文件路径
    """
    try:
        # 确保UTF-8编码环境
        ensure_utf8_encoding()
        
        if output_dir is None:
            output_dir = xlsx_path.parent
        
        soffice_path = find_soffice_executable()
        
        # 标准化输入文件路径
        xlsx_path_str = str(xlsx_path)
        output_dir_str = str(output_dir)
        
        # 增强转换参数，移除强制分页设置，保持Excel原有格式
        cmd = [
            soffice_path, 
            "--headless",
            "--infilter=Calc8",
            "--convert-to", "pdf:calc_pdf_Export:{'EmbedComplexScriptFonts':true,'EmbedFonts':true,'ExportNotes':false}",
            "--outdir", output_dir_str,
            xlsx_path_str
        ]
        
        # 设置环境变量确保UTF-8编码
        env = os.environ.copy()
        env.update({
            'LC_ALL': 'C.UTF-8',
            'LANG': 'C.UTF-8',
            'LC_CTYPE': 'C.UTF-8'
        })
        
        result = subprocess.run(cmd, check=True, env=env, 
                               capture_output=True, text=True, encoding='utf-8', errors='ignore')
        
        # 生成PDF文件路径
        pdf_path = output_dir / xlsx_path.with_suffix('.pdf').name
        
        # 检查生成的文件是否存在，如果不存在尝试查找可能的编码问题
        if not pdf_path.exists():
            # 列出输出目录中的所有PDF文件
            pdf_files = list(output_dir.glob("*.pdf"))
            if pdf_files:
                # 找到最新的PDF文件
                latest_pdf = max(pdf_files, key=lambda p: p.stat().st_mtime)
                # 如果文件名有编码问题，重命名为正确的文件名
                expected_name = xlsx_path.stem + '.pdf'
                if latest_pdf.name != expected_name:
                    correct_pdf_path = output_dir / expected_name
                    latest_pdf.rename(correct_pdf_path)
                    pdf_path = correct_pdf_path
                    logger.info(f"已重命名PDF文件: {latest_pdf} -> {pdf_path}")
                else:
                    pdf_path = latest_pdf
        
        logger.info(f"PDF文件生成完成: {pdf_path}")
        return pdf_path
        
    except Exception as e:
        logger.error(f"Excel转PDF时出错: {str(e)}")
        raise


def get_poppler_path() -> str:
    """
    获取Poppler工具路径，仅Windows需要显式指定
    
    Returns:
        str: Poppler路径，非Windows系统返回None
    """
    if platform.system() == "Windows":
        # Windows系统需要指定Poppler路径
        return r"D:\Program Files\poppler-24.08.0\Library\bin"
    return None


def convert_pdf_to_png(pdf_path: Path, output_path: Path = None, dpi: int = 200) -> Path:
    """
    将PDF文件转换为PNG图片（仅转换第一页）
    
    Args:
        pdf_path: PDF文件路径
        output_path: 输出PNG文件路径，如果不指定则使用PDF文件名
        dpi: 图片分辨率，默认200
        
    Returns:
        Path: 生成的PNG图片路径
    """
    try:
        # 确保UTF-8编码环境
        ensure_utf8_encoding()
        
        # 标准化PDF文件路径
        pdf_path_normalized = Path(normalize_filename(str(pdf_path)))
        if not pdf_path_normalized.exists() and pdf_path.exists():
            pdf_path_normalized = pdf_path
        
        if output_path is None:
            output_path = pdf_path_normalized.with_suffix('.png')
        
        poppler_path = get_poppler_path()
        
        # 设置环境变量确保UTF-8编码
        original_env = os.environ.copy()
        os.environ.update({
            'LC_ALL': 'C.UTF-8',
            'LANG': 'C.UTF-8',
            'LC_CTYPE': 'C.UTF-8'
        })
        
        try:
            pages = convert_from_path(
                str(pdf_path_normalized), 
                dpi=dpi,
                poppler_path=poppler_path
            )
            
            # 只保存第一页
            pages[0].save(str(output_path), "PNG")
            logger.info(f"PNG图片生成完成: {output_path}")
            return output_path
        finally:
            # 恢复原环境变量
            os.environ.clear()
            os.environ.update(original_env)
        
    except Exception as e:
        logger.error(f"PDF转PNG时出错: {str(e)}")
        # 如果PDF文件路径有编码问题，尝试查找并使用正确的文件
        try:
            pdf_dir = pdf_path.parent
            pdf_files = list(pdf_dir.glob("*.pdf"))
            if pdf_files:
                # 查找匹配的PDF文件（基于文件大小和修改时间）
                target_pdf = None
                for pdf_file in pdf_files:
                    if pdf_file.stat().st_size > 0:  # 确保文件不为空
                        target_pdf = pdf_file
                        break
                
                if target_pdf:
                    logger.info(f"尝试使用找到的PDF文件: {target_pdf}")
                    pages = convert_from_path(str(target_pdf), dpi=dpi, poppler_path=get_poppler_path())
                    pages[0].save(str(output_path), "PNG")
                    logger.info(f"PNG图片生成完成: {output_path}")
                    return output_path
        except Exception as fallback_error:
            logger.error(f"备用转换方案也失败: {fallback_error}")
        
        raise


def ensure_directory_exists(directory_path: str) -> Path:
    """
    确保目录存在，如果不存在则创建
    
    Args:
        directory_path: 目录路径
        
    Returns:
        Path: 目录路径对象
    """
    dir_path = Path(directory_path)
    dir_path.mkdir(parents=True, exist_ok=True)
    return dir_path 