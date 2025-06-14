import os, platform, shutil, subprocess
from pathlib import Path
from pdf2image import convert_from_path
from openpyxl import load_workbook
from openpyxl.styles import Side, Border

def set_landscape(xlsx_path, sheetname=None):
    """把指定 Excel 的 sheet 设置为横向打印、A4纸、一页宽、内容居中、自动打印区域，并修复合并标题右边框问题"""
    wb = load_workbook(xlsx_path)
    ws = wb[sheetname] if sheetname else wb.active

    # 横向 & A4
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_A4

    # 一页宽（不限行数）
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    # 内容水平居中
    ws.print_options.horizontalCentered = True

    # 自动设置打印区域（所有已用单元格）
    min_row = ws.min_row
    max_row = ws.max_row
    min_col = ws.min_column
    max_col = ws.max_column
    start_cell = ws.cell(row=min_row, column=min_col).coordinate
    end_cell = ws.cell(row=max_row, column=max_col).coordinate
    ws.print_area = f"{start_cell}:{end_cell}"

    # 修复A1:L1标题区右侧边框加粗问题
    # 如果你的标题区不是A1:L1，请相应修改
    right_title_cell = ws['L1']
    # 用现有边框信息，右侧设为medium
    right_title_cell.border = Border(
        left=right_title_cell.border.left,
        right=Side(style='medium'),
        top=right_title_cell.border.top,
        bottom=right_title_cell.border.bottom
    )

    wb.save(xlsx_path)

def find_soffice() -> str:
    """动态寻找 soffice"""
    sys = platform.system()
    candidates = []
    if sys == "Windows":
        candidates += [
            r"C:\Program Files\LibreOffice\program\soffice.exe",
            r"C:\Program Files (x86)\LibreOffice\program\soffice.exe",
        ]
    elif sys == "Darwin":      # macOS
        candidates += [
            "/Applications/LibreOffice.app/Contents/MacOS/soffice",
            shutil.which("soffice"),
        ]
    else:                       # Linux
        candidates += [shutil.which("soffice"), shutil.which("libreoffice")]
    for c in candidates:
        if c and Path(c).exists():
            return c
    raise FileNotFoundError("无法找到 LibreOffice 的 soffice，可执行文件未安装或不在 PATH")

def xlsx_to_pdf(xlsx_path: Path, pdf_path: Path):
    subprocess.run([
        find_soffice(), "--headless",
        "--convert-to", "pdf",
        "--outdir", str(pdf_path.parent),
        str(xlsx_path)
    ], check=True)

def get_poppler_path():
    """仅 Windows 需显式 poppler_path"""
    return (r"D:\Program Files\poppler-24.08.0\Library\bin"
            if platform.system() == "Windows" else None)

def pdf_to_png(pdf_path: Path, png_path: Path):
    pages = convert_from_path(str(pdf_path), dpi=200,
                              poppler_path=get_poppler_path())
    pages[0].save(str(png_path), "PNG")

if __name__ == "__main__":
    # 路径配置
    xlsx_file = Path("市场销售数据_2025-06-12.xlsx")
    pdf_file = xlsx_file.with_suffix('.pdf')
    png_file = xlsx_file.with_suffix('.png')

    # 1. 先设置为横向打印（纸张A4+全内容+居中）
    set_landscape(str(xlsx_file), sheetname="市场日报")

    # 2. 转PDF
    xlsx_to_pdf(xlsx_file, pdf_file)
    print("已生成PDF:", pdf_file)

    # 3. PDF转PNG
    pdf_to_png(pdf_file, png_file)
    print("已生成PNG:", png_file)
