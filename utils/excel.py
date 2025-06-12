import json
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00
from openpyxl.utils import get_column_letter
from datetime import datetime


def convert_percentage_to_decimal(percentage_str):
    """
    将百分比字符串转换为小数
    例如："23.1%" -> 0.231
    """
    if isinstance(percentage_str, str) and percentage_str.endswith('%'):
        return float(percentage_str[:-1]) / 100
    return percentage_str


def export_json_to_excel(json_data, report_date, filename=None):
    """
    将JSON数据导出为Excel文件，按照指定格式和样式
    
    json_data: 数据集
    report_date: 报告日期，格式“YYYY-MM-DD”
    filename: 输出文件名，例如"市场销售数据_6月10日.xlsx"
    """

    # 创建工作簿和工作表
    wb = Workbook()
    ws = wb.active
    ws.title = "市场日报"

    # 设置列宽
    column_widths = {
        'A': 12.4, 'B': 5.2, 'C': 14.68, 'D': 9.2, 'E': 13.2, 'F': 10,
        'G': 10, 'H': 12.24, 'I': 11.2, 'J': 11.2, 'K': 10.24, 'L': 10.24
    }
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # 定义样式
    # 边框样式
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    medium_border = Border(
        left=Side(style='medium'),
        right=Side(style='medium'),
        top=Side(style='medium'),
        bottom=Side(style='medium')
    )

    # 字体样式
    title_font = Font(name='Microsoft YaHei', size=16, bold=True, color='FFFFFF')
    header_font = Font(name='Microsoft YaHei', size=11, bold=True)
    market_font = Font(name='Microsoft YaHei', size=11, bold=True)
    data_font = Font(name='Microsoft YaHei', size=10)
    summary_font = Font(name='Microsoft YaHei', size=10, bold=True)

    # 填充色
    title_fill = PatternFill(start_color='333F4F', end_color='333F4F', fill_type='solid')

    # 对齐方式
    center_alignment = Alignment(horizontal='center', vertical='center')
    left_alignment = Alignment(horizontal='left', vertical='center')

    current_row = 1

    # 1. 标题区
    dt = datetime.strptime(report_date, "%Y-%m-%d")
    # 汉字部分，先分开
    title_head = ' '.join("市场销售日报")
    # 日期部分，用空格手动拼接
    date_str = f"{dt.month} 月 {dt.day} 日"
    title = f"{title_head}  {date_str}"
    ws.merge_cells('A1:L1')
    title_cell = ws['A1']
    title_cell.value = title
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.alignment = center_alignment
    title_cell.border = medium_border
    ws.row_dimensions[current_row].height = 33
    current_row += 1

    # 2. 表头区 (A2:L2)
    headers = ['区域', '序号', '名称', '车辆配置', '当日销售', '当日车均', '当日车次',
               '月目标', '月累计', '累计达成率', '累计车均', '累计车次']

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=current_row, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = medium_border

    ws.row_dimensions[current_row].height = 22.5
    current_row += 1

    # 3. 数据主体区
    for market_data in json_data['markets']:
        market_name = market_data['name']
        warehouses = market_data['warehouses']

        # 计算市场分组需要合并的行数（仓库数量 + 1个小计行）
        market_rows = len(warehouses) + 1
        market_start_row = current_row

        # 市场分组行 - 先写入市场名称
        market_cell = ws.cell(row=current_row, column=1)
        market_cell.value = f'📍{market_name}'
        market_cell.font = market_font
        market_cell.alignment = center_alignment
        market_cell.border = medium_border

        # 仓库数据行
        for warehouse in warehouses:
            # 序号 (id)
            ws.cell(row=current_row, column=2, value=warehouse['id'])
            # 名称 (name)
            ws.cell(row=current_row, column=3, value=warehouse['name'])
            # 车辆配置 (car_count)
            ws.cell(row=current_row, column=4, value=warehouse['car_count'])
            # 当日销售 (daily_revenue)
            ws.cell(row=current_row, column=5, value=warehouse['daily_revenue'])
            # 当日车均 (daily_avg_revenue_cart)
            ws.cell(row=current_row, column=6, value=warehouse['daily_avg_revenue_cart'])
            # 当日车次 (daily_cart_count)
            ws.cell(row=current_row, column=7, value=warehouse['daily_cart_count'])
            # 月目标 (target_income)
            ws.cell(row=current_row, column=8, value=warehouse['target_income'])
            # 月累计 (actual_income)
            ws.cell(row=current_row, column=9, value=warehouse['actual_income'])
            # 累计达成率 (ach_rate)
            ach_rate_cell = ws.cell(row=current_row, column=10, value=convert_percentage_to_decimal(warehouse['ach_rate']))
            ach_rate_cell.number_format = '0.0%'
            # 累计车均 (per_car_income)
            ws.cell(row=current_row, column=11, value=warehouse['per_car_income'])
            # 累计车次 (sold_car_count)
            ws.cell(row=current_row, column=12, value=warehouse['sold_car_count'])

            # 设置数据行样式
            for col in range(2, 13):
                cell = ws.cell(row=current_row, column=col)
                cell.font = data_font
                cell.alignment = left_alignment if col == 3 else center_alignment
                cell.border = thin_border

            ws.row_dimensions[current_row].height = 22.5
            current_row += 1

        # 合并市场分组单元格
        if market_rows > 1:
            ws.merge_cells(f'A{market_start_row}:A{market_start_row + market_rows - 1}')

        # 小计行
        summary = market_data['summary']

        # 合并B2:C列作为"小计"
        ws.merge_cells(f'B{current_row}:C{current_row}')
        subtotal_cell = ws.cell(row=current_row, column=2)
        subtotal_cell.value = '小计'
        subtotal_cell.font = summary_font
        subtotal_cell.alignment = center_alignment
        subtotal_cell.border = thin_border

        # 小计数据
        ws.cell(row=current_row, column=4, value=summary['car_count'])
        ws.cell(row=current_row, column=5, value=summary['daily_revenue'])
        ws.cell(row=current_row, column=6, value=summary['daily_avg_revenue_cart'])
        ws.cell(row=current_row, column=7, value=summary['daily_cart_count'])
        ws.cell(row=current_row, column=8, value=summary['target_income'])
        ws.cell(row=current_row, column=9, value=summary['actual_income'])
        ach_rate_cell = ws.cell(row=current_row, column=10, value=convert_percentage_to_decimal(summary['ach_rate']))
        ach_rate_cell.number_format = '0.0%'
        ws.cell(row=current_row, column=11, value=summary['per_car_income'])
        ws.cell(row=current_row, column=12, value=summary['sold_car_count'])

        # 设置小计行样式
        for col in range(4, 13):
            cell = ws.cell(row=current_row, column=col)
            cell.font = summary_font
            cell.alignment = center_alignment
            cell.border = thin_border

        ws.row_dimensions[current_row].height = 22.5
        current_row += 1

        # 为整个市场区域设置medium边框
        market_end_row = current_row - 1
        # 设置区域外边框
        for row in range(market_start_row, market_end_row + 1):
            for col in range(1, 13):
                cell = ws.cell(row=row, column=col)
                # 保持原有边框，但在区域边界添加medium边框
                current_border = cell.border
                new_border = Border(
                    left=Side(style='medium') if col == 1 else current_border.left,
                    right=Side(style='medium') if col == 12 else current_border.right,
                    top=Side(style='medium') if row == market_start_row else current_border.top,
                    bottom=Side(style='medium') if row == market_end_row else current_border.bottom
                )
                cell.border = new_border

    # 4. 总计行
    total_data = json_data['total']

    # 合并A列到C列作为"总计"，先设置所有单元格的边框
    for col in range(1, 4):
        cell = ws.cell(row=current_row, column=col)
        cell.border = medium_border
        if col == 1:
            cell.value = '总计'
            cell.font = summary_font
            cell.alignment = center_alignment
    
    # 合并单元格
    ws.merge_cells(f'A{current_row}:C{current_row}')

    # 总计数据
    ws.cell(row=current_row, column=4, value=total_data['car_count'])
    ws.cell(row=current_row, column=5, value=total_data['daily_revenue'])
    ws.cell(row=current_row, column=6, value=total_data['daily_avg_revenue_cart'])
    ws.cell(row=current_row, column=7, value=total_data['daily_cart_count'])
    ws.cell(row=current_row, column=8, value=total_data['target_income'])
    ws.cell(row=current_row, column=9, value=total_data['actual_income'])
    ach_rate_cell = ws.cell(row=current_row, column=10, value=convert_percentage_to_decimal(total_data['ach_rate']))
    ach_rate_cell.number_format = '0.0%'
    ws.cell(row=current_row, column=11, value=total_data['per_car_income'])
    ws.cell(row=current_row, column=12, value=total_data['sold_car_count'])

    # 设置总计行样式
    for col in range(4, 13):
        cell = ws.cell(row=current_row, column=col)
        cell.font = summary_font
        cell.alignment = center_alignment
        cell.border = medium_border

    ws.row_dimensions[current_row].height = 33

    # 保存文件
    wb.save(filename)
    print(f"Excel文件已保存为: {filename}")


# 不再需要这里的使用示例，将直接从daily_sales.py中调用
if __name__ == "__main__":
    with open("data/daily_sales.json", "r", encoding="utf-8") as f:
        json_data = json.load(f)
    export_json_to_excel(json_data)